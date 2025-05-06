import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import tkinter as ttk
import customtkinter as ctk
from tkinter import ttk
from tkinter import filedialog, messagebox
from views.crop_anh_view import CropperWindow
import cv2
from resources.models_ai.nhan_dien import DigitRecognizer
from resources.models_ai.xu_ly_anh import ImageProcessor
from controllers.diem_controller import DiemController

class NhapDiemChiTietFrame(ctk.CTkFrame):
    def __init__(self, master=None, app=None, ma_mon=None, ten_mon=None, so_luong=None, ma_lop=None, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(fg_color="#ffffff")
        self.app = app
        self.ma_mon = ma_mon
        self.ten_mon = ten_mon
        self.so_luong = so_luong
        self.ma_lop = ma_lop
        self.controller = DiemController()

        content_frame = ctk.CTkFrame(self, fg_color="#ffffff")
        content_frame.pack(fill="both", expand=True)

        # Tiêu đề
        header_frame = ctk.CTkFrame(content_frame, fg_color="#646765", height=100)
        header_frame.pack(fill="x")

        label_title = ctk.CTkLabel(
            header_frame,
            text="Nhập điểm cuối kỳ",
            font=("Verdana", 18, "bold"),
            text_color="#ffffff"
        )
        label_title.pack(pady=20)

        # Thông tin môn học
        info_frame = ctk.CTkFrame(content_frame, fg_color="white")
        info_frame.pack(fill="x", padx=20, pady=10)

        self.label_mon = ctk.CTkLabel(info_frame, text=f"Môn: {self.ten_mon}", font=("Verdana", 13))
        self.label_mon.grid(row=0, column=0, sticky="w", padx=10)

        self.label_so_luong = ctk.CTkLabel(info_frame, text=f"Số lượng: {self.so_luong}", font=("Verdana", 13))
        self.label_so_luong.grid(row=0, column=1, sticky="w", padx=30)

        self.btn_xuat_excel = ctk.CTkButton(
            info_frame,  # Place buttons in info_frame
            text="Xuất Excel",
            fg_color="#4CAF50",
            text_color="white",
            font=("Verdana", 12, "bold"),
            command=self.export_excel
        )
        self.btn_xuat_excel.grid(row=0, column=2, padx=10)

        self.btn_tai_anh = ctk.CTkButton(
            info_frame,  # Place buttons in info_frame
            text="Tải ảnh lên",
            fg_color="#904fd2",
            text_color="white",
            font=("Verdana", 12, "bold"),
            command=self.upload_diem_cuoi_ky
        )
        self.btn_tai_anh.grid(row=0, column=3, padx=10)

        self.btn_sua = ctk.CTkButton(
            info_frame,  # Place buttons in info_frame
            text="Sửa",
            fg_color="#FFA500",  # Orange color
            text_color="white",
            font=("Verdana", 12, "bold"),
            command=self.edit_diem  # Add edit functionality later
        )
        self.btn_sua.grid(row=0, column=4, padx=10)

        # Bảng điểm sinh viên
        tree_frame = ctk.CTkFrame(content_frame, fg_color="white")
        tree_frame.pack(fill="both", expand=True, padx=20, pady=10)

        style = ttk.Style()
        style.configure("Treeview",
                        background="#f5f5f5",
                        foreground="black",
                        rowheight=30,
                        fieldbackground="lightgray")
        style.configure("Treeview.Heading",
                        font=("Arial", 12, "bold"),
                        background="#3084ee",
                        foreground="black")
        style.map("Treeview",
                  background=[("selected", "#4CAF50")],
                  foreground=[("selected", "white")])

        self.tree = ttk.Treeview(tree_frame, columns=("mssv", "ten_sv", "diem"), show="headings", style="Treeview")
        self.tree.heading("mssv", text="MSSV")
        self.tree.heading("ten_sv", text="Tên sinh viên")
        self.tree.heading("diem", text="Điểm cuối kỳ")

        self.tree.column("mssv", width=100, anchor="center")
        self.tree.column("ten_sv", width=200, anchor="center")
        self.tree.column("diem", width=120, anchor="center")

        self.tree.pack(fill="both", expand=True)

        self.load_data()

    def load_data(self):
        data = self.controller.get_diem_cuoi_ky_by_ma_mon(self.ma_lop)
        print(data)
        if data:
            for row in self.tree.get_children():
                self.tree.delete(row)
            for sv in data:
                mssv = sv.get("mssv", "")
                ho_ten = sv.get("ho_ten", "")
                diem = sv.get("diem_cuoi_ky", "") if sv.get("diem_cuoi_ky") is not None else ""
                self.tree.insert('', 'end', values=(mssv, ho_ten, diem))
        else:
            messagebox.showinfo("Thông báo", "Không có dữ liệu để hiển thị.")

    def export_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Lưu bảng điểm đơn giản"
        )
        if not file_path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Bảng Điểm"

        # Định nghĩa font Times New Roman, size 14
        default_font = Font(name="Times New Roman", size=14)
        bold_font = Font(name="Times New Roman", size=14, bold=True)

        # Căn giữa
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Định nghĩa đường viền
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Ghi tiêu đề cột
        headers = ["MSSV", "Tên Sinh Viên", "Điểm"]
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment

        # Ghi dữ liệu từng dòng
        for row_index, item in enumerate(self.tree.get_children(), start=2):
            values = self.tree.item(item, "values")
            try:
                mssv = int(values[0])  # ép kiểu để không bị cảnh báo số lưu dưới dạng text
            except:
                mssv = values[0]

            ho_ten = values[1]
            diem = ""  # để trống

            row_data = [mssv, ho_ten, diem]
            for col_index, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.font = default_font
                cell.alignment = center_alignment
                cell.border = thin_border

        # Auto fit độ rộng cột
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 4
            sheet.column_dimensions[column].width = adjusted_width

        try:
            workbook.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất file Excel:\n{file_path}")
            os.startfile(file_path)  # Mở file Excel sau khi lưu
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu file Excel:\n{str(e)}")

    def upload_diem_cuoi_ky(self):
        file_path = filedialog.askopenfilename(
            title="Chọn file ảnh điểm cuối kỳ",
            filetypes=[("Image Files", "*.png *.jpg *.jpeg *.gif *.bmp")]
        )
        if not file_path:
            return

        def handle_cropped_image(cropped_path):
            print("Đã cắt ảnh:", cropped_path)
            messagebox.showinfo("Tải ảnh thành công", f"Đã chọn file:\n{file_path}")

            # Gọi luôn xử lý, không cần chọn cột
            self.handle_column_button_cuoi_ky(cropped_path)

        # Gọi cửa sổ crop ảnh
        CropperWindow(self, file_path, handle_cropped_image)

    def handle_column_button_cuoi_ky(self, file_path):
        ten_cot_diem = "Điểm Cuối Kỳ"
        messagebox.showinfo("Xử lý ảnh", f"Đã chọn ảnh:\n{file_path}\nGán điểm cho cột: {ten_cot_diem}")

        image = cv2.imread(file_path)
        recognizer = DigitRecognizer()
        processor = ImageProcessor(image, recognizer)
        processor.extract_and_recognize()
        result_kq = processor.result_kq  # [(mssv, diem), ...]
        print(result_kq)

        self.show_edit_popup(result_kq, ten_cot_diem)

    def show_edit_popup(self, result_kq, ten_cot_diem):
        # Bản đồ từ MSSV OCR → điểm
        map_ocr = dict(result_kq)

        # Danh sách sinh viên trong lớp
        matched_any = False
        sinh_vien_rows = []

        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            mssv = values[0]
            ten_sv = values[1]
            diem = map_ocr.get(mssv, "")  # nếu không khớp thì để trống
            if mssv in map_ocr:
                matched_any = True
            sinh_vien_rows.append((mssv, ten_sv, diem))

        if not matched_any:
            messagebox.showwarning("Không tìm thấy",
                                   "❌ Không tìm thấy MSSV nào phù hợp trong ảnh bảng điểm.\nVui lòng kiểm tra lại ảnh.")
            return

        # === Tạo cửa sổ popup ===
        popup = ctk.CTkToplevel(self)
        popup.title(f"Chỉnh sửa điểm - {ten_cot_diem}")
        popup.geometry("600x400")
        popup.lift()
        popup.attributes("-topmost", True)

        frame = ctk.CTkFrame(popup)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        columns = ("MSSV", "Tên SV", "Điểm")
        tree = ttk.Treeview(frame, columns=columns, show="headings")
        tree.pack(fill="both", expand=True)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=180)

        for mssv, ten_sv, diem in sinh_vien_rows:
            tree.insert("", "end", values=(mssv, ten_sv, diem))

        # === Cho phép chỉnh sửa cột điểm ===
        def on_double_click(event):
            item_id = tree.identify_row(event.y)
            col = tree.identify_column(event.x)

            if col != "#3":
                return

            x, y, width, height = tree.bbox(item_id, column=col)
            value = tree.set(item_id, column=col)

            entry = ttk.Entry(tree)
            entry.place(x=x, y=y, width=width, height=height)
            entry.insert(0, value)
            entry.focus()

            def on_focus_out(event):
                new_value = entry.get()
                tree.set(item_id, column=col, value=new_value)
                entry.destroy()

            entry.bind("<FocusOut>", on_focus_out)
            entry.bind("<Return>", lambda e: on_focus_out(e))

        tree.bind("<Double-1>", on_double_click)

        def save_data():
            for item in tree.get_children():
                mssv, _, diem = tree.item(item, "values")
                if diem.strip() == "":
                    continue

                try:
                    diem_float = float(diem)
                    if diem_float < 0 or diem_float > 10:
                        raise ValueError
                except ValueError:
                    messagebox.showerror("Lỗi", f"Điểm không hợp lệ cho MSSV {mssv}. Hãy kiểm tra lại.")
                    return

                self.controller.update(mssv, self.ma_lop, diem)

        save_btn = ctk.CTkButton(popup, text="Lưu dữ liệu vào bảng điểm", command=save_data)
        save_btn.pack(pady=10)

    def edit_diem(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Chọn dòng", "Vui lòng chọn một sinh viên để sửa.")
            return

        values = self.tree.item(selected_item, "values")
        if not values:
            return

        # Tạo popup sửa điểm
        popup = ctk.CTkToplevel(self)
        popup.title("Sửa điểm cuối kỳ")  # Changed title
        popup.geometry("400x400")  # Smaller size

        # Hiển thị MSSV (không cho sửa)
        ctk.CTkLabel(popup, text="MSSV").pack(anchor="w", padx=10, pady=5)
        ctk.CTkLabel(popup, text=values[0]).pack(anchor="w", padx=10, pady=5)

        # Hiển thị Họ tên (không cho sửa)
        ctk.CTkLabel(popup, text="Họ tên").pack(anchor="w", padx=10, pady=5)
        ctk.CTkLabel(popup, text=values[1]).pack(anchor="w", padx=10, pady=5)

        # Điểm cuối kỳ (cho sửa)
        ctk.CTkLabel(popup, text="Điểm cuối kỳ").pack(anchor="w", padx=10, pady=5)
        diem_var = ctk.StringVar(value=values[2] if values[2] != "" else "0")  # Default to 0 if empty

        entry_diem = ctk.CTkEntry(popup, textvariable=diem_var)
        entry_diem.pack(fill="x", padx=10, pady=5)

        def save_edits():
            mssv = values[0]
            diem_str = diem_var.get()

            try:
                diem = float(diem_str)
                if not (0 <= diem <= 10):
                    raise ValueError  # Kiểm tra khoảng điểm
            except ValueError:
                messagebox.showerror("Lỗi", "Điểm không hợp lệ. Vui lòng nhập số từ 0 đến 10.")
                return  # Dừng lưu nếu có lỗi

            self.controller.update(mssv, self.ma_lop, diem)  # Corrected controller function name

            self.load_data()  # Refresh the treeview
            popup.destroy()
            messagebox.showinfo("Thành công", "Đã cập nhật điểm.")

        save_btn = ctk.CTkButton(popup, text="Lưu", command=save_edits)
        save_btn.pack(pady=10)






