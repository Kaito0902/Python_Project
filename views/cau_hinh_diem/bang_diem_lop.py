import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import cv2
from resources.models_ai.nhan_dien import DigitRecognizer
from resources.models_ai.xu_ly_anh import ImageProcessor
import customtkinter as ctk
from tkinter import ttk, messagebox
from tkinter import filedialog
from controllers.chi_tiet_diem_controller import ChiTietDiemController
from controllers.lop_controller import LopController
from controllers.sinh_vien_controller import SinhVienController
from controllers.cau_hinh_diem_controller import CauHinhDiemController
from controllers.mon_hoc_controller import MonHocController
from controllers.diem_controller import DiemController
from views.crop_anh_view import CropperWindow


class BangDiemLop(ctk.CTkFrame):
    def __init__(self, parent, ma_lop=None):
        super().__init__(parent, corner_radius=15, fg_color="white")
        self.pack(fill="both", expand=True, padx=10, pady=10)
        self.ma_lop = ma_lop
        self.controller = ChiTietDiemController()
        self.lopcontroller = LopController()
        self.svcontroller = SinhVienController()
        self.chdcontroller = CauHinhDiemController()
        self.moncontroller = MonHocController()
        self.diemcontroller = DiemController()

        self.cot_co_dinh = ["MSSV", "Tên Sinh Viên", "Điểm Kiểm Tra"]
        self.cot_diem_list = []
        self.id_cot_diem = []
        self.ds = []

        header_frame = ctk.CTkFrame(self, fg_color="#646765", height=100)
        header_frame.pack(fill="x")

        label_title = ctk.CTkLabel(header_frame, text="Bảng Điểm Lớp", font=("Verdana", 18, "bold"),
                                   text_color="#ffffff")
        label_title.pack(pady=20)

        self.info_frame = ctk.CTkFrame(self)
        self.info_frame.pack(fill="x", pady=5, padx=10)

        labels = ["Mã lớp:", "Môn học:", "Số lượng SV:"]
        values = [self.ma_lop, "", ""]
        self.info_labels = []

        for i, (label_text, value_text) in enumerate(zip(labels, values)):
            label = ctk.CTkLabel(self.info_frame, text=label_text, width=15, anchor="w")
            label.grid(row=0, column=2 * i, padx=(0, 5), pady=2, sticky="w")
            value_label = ctk.CTkLabel(self.info_frame, text=value_text, anchor="w")
            value_label.grid(row=0, column=2 * i + 1, padx=(0, 20), pady=2, sticky="w")
            self.info_labels.append((label, value_label))

        button_width = 100
        button_height = 30
        button_font = ("Arial", 12)

        self.upload_button = ctk.CTkButton(
            self.info_frame, text="Tải ảnh lên", command=self.upload_image,
            width=button_width, height=button_height, font=button_font
        )
        self.upload_button.grid(row=0, column=6, padx=5, pady=5, sticky="e")

        self.export_button = ctk.CTkButton(
            self.info_frame, text="Xuất Excel", command=self.export_to_excel,
            width=button_width, height=button_height, font=button_font
        )
        self.export_button.grid(row=0, column=7, padx=5, pady=5, sticky="e")

        self.edit_button = ctk.CTkButton(
            self.info_frame, text="Sửa điểm", command=self.edit_selected_row,
            width=button_width, height=button_height, font=button_font
        )
        self.edit_button.grid(row=0, column=8, padx=5, pady=5, sticky="e")

        # Treeview
        self.tree_frame = ctk.CTkFrame(self)
        self.tree_frame.pack(fill="both", expand=True)
        self.load_data()
        self.tinh_diem_kiem_tra()

    def create_treeview(self):

        for widget in self.tree_frame.winfo_children():
            widget.destroy()

        index_diem_kiem_tra = self.cot_co_dinh.index("Điểm Kiểm Tra")

        cot_truoc_diem_kt = self.cot_co_dinh[:index_diem_kiem_tra]
        cot_sau_diem_kt = self.cot_co_dinh[index_diem_kiem_tra:]

        columns = cot_truoc_diem_kt + [ten for ten, ts in self.cot_diem_list] + cot_sau_diem_kt

        style = ttk.Style()
        style.configure("Treeview", background="#f5f5f5", foreground="black", rowheight=30, fieldbackground="lightgray")
        style.configure("Treeview.Heading", font=("Arial", 12, "bold"), background="#3084ee", foreground="black")
        style.map("Treeview", background=[("selected", "#4CAF50")], foreground=[("selected", "white")])

        self.tree = ttk.Treeview(self.tree_frame, columns=columns, show="headings", style="Treeview")
        self.tree.pack(fill="both", expand=True)

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120)

    def fill_tree(self, chi_tiet_diem_list):
        for row in self.tree.get_children():
            self.tree.delete(row)

        for sv_data in chi_tiet_diem_list:
            mssv = sv_data.get("mssv", "")
            ho_ten = sv_data.get("ho_ten", "")
            row = [mssv, ho_ten]
            for ten_cot_diem, _ in self.cot_diem_list:
                row.append(sv_data.get(ten_cot_diem, ""))
            row.append("")  # Cột "Điểm Kiểm Tra" mặc định
            self.tree.insert("", "end", values=row)

    def load_data(self):
        self.lopcontroller = LopController()
        lop_info = self.lopcontroller.get_by_ma_lop(self.ma_lop)
        ma_mon = lop_info.get("ma_mon", "") if lop_info else ""
        ten_mon = self.moncontroller.select_by_id(ma_mon)[0]["ten_mon"]

        # Lấy danh sách sinh viên
        ds_sinh_vien = self.lopcontroller.get_students_in_class(self.ma_lop)
        so_luong_sv = len(ds_sinh_vien)

        # Cập nhật label thông tin
        self.info_labels[1][1].configure(text=ten_mon)  # Môn học
        self.info_labels[2][1].configure(text=str(so_luong_sv))  # Số lượng SV

        # Lấy danh sách cột điểm theo cấu hình
        self.chdcontroller = CauHinhDiemController()
        ds_cot_diem = self.chdcontroller.select_all(self.ma_lop)
        self.id_cot_diem = [[cot["id"]] for cot in ds_cot_diem]
        self.cot_diem_list = [(cot["ten_cot_diem"], cot["trong_so"]) for cot in ds_cot_diem]
        self.ds = ds_cot_diem

        # Tạo lại bảng với các cột
        self.create_treeview()

        # Lấy chi tiết điểm đã có (nếu có)
        self.controller = ChiTietDiemController()
        chi_tiet_diem_list = self.controller.select_all(self.ma_lop)

        # Duyệt sinh viên
        for sv in ds_sinh_vien:
            mssv = sv["mssv"]
            ten_sv = sv["ho_ten"]
            diem_dict = {cot["ten_cot_diem"]: "" for cot in ds_cot_diem}  # mặc định chưa có điểm

            # Gán điểm nếu có
            for diem in chi_tiet_diem_list:
                if diem["mssv"] == mssv:
                    ten_cot = diem["ten_cot_diem"]
                    diem_dict[ten_cot] = diem["diem"]

            # Chuẩn bị dòng dữ liệu
            values = [mssv, ten_sv] + [diem_dict[cot[0]] for cot in self.cot_diem_list] + [
                ""]  # "" cho cột "Điểm kiểm tra"
            self.tree.insert("", "end", values=values)

    # def add_new_column(self, ten_cot_diem, trong_so):
    #     self.cot_diem_list.append((ten_cot_diem, trong_so))
    #     self.create_treeview()
    #
    # def remove_column(self, ten_cot_diem):
    #     self.cot_diem_list = [(ten, ts) for ten, ts in self.cot_diem_list if ten != ten_cot_diem]
    #     self.create_treeview()

    def refresh_columns_and_data(self):
        self.clear_table()
        self.load_data()

    def clear_table(self):
        if hasattr(self, 'tree') and self.tree:
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.tree.destroy()

    def upload_image(self):
        file_path = filedialog.askopenfilename(
            title="Chọn file ảnh",
            filetypes=[("Image Files", "*.png *.jpg *.jpeg *.gif *.bmp")]
        )
        if not file_path:
            return

        def handle_cropped_image(cropped_path):
            print("Đã cắt ảnh:", cropped_path)
            messagebox.showinfo("Tải ảnh thành công", f"Đã chọn file:\n{file_path}")

            # Tạo cửa sổ popup sau khi đã crop ảnh xong
            popup = ctk.CTkToplevel(self)
            popup.title("Chọn cột điểm")
            popup.geometry("400x300")
            popup.lift()
            popup.attributes("-topmost", True)

            label = ctk.CTkLabel(popup, text="Chọn cột điểm để gán kết quả từ ảnh:", font=("Arial", 14))
            label.pack(pady=10)

            for ten_cot, _ in self.cot_diem_list:
                if ten_cot == "Điểm Kiểm Tra":
                    continue
                btn = ctk.CTkButton(popup, text=ten_cot, width=200,
                                    command=lambda t=ten_cot: self.handle_column_button(t, cropped_path, popup))
                btn.pack(pady=5)

        # Gọi cửa sổ crop ảnh
        CropperWindow(self, file_path, handle_cropped_image)

    def handle_column_button(self, ten_cot_diem, file_path, popup_window):
        popup_window.destroy()
        messagebox.showinfo("Xử lý ảnh", f"Đã chọn ảnh:\n{file_path}\nGán điểm cho cột: {ten_cot_diem}")

        image = cv2.imread(file_path)
        recognizer = DigitRecognizer()
        processor = ImageProcessor(image, recognizer)
        processor.extract_and_recognize()
        result_kq = processor.result_kq  # [(mssv, diem), ...]

        self.show_edit_popup(result_kq, ten_cot_diem)

    def show_edit_popup(self, result_kq, ten_cot_diem):
        # Bản đồ từ MSSV OCR → điểm
        map_ocr = dict(result_kq)

        # Danh sách sinh viên trong lớp
        matched_any = False
        sinh_vien_rows = []

        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            mssv = values[0]
            ten_sv = values[1]
            diem = map_ocr.get(mssv, "")  # nếu không khớp thì để trống
            if mssv in map_ocr:
                matched_any = True
            sinh_vien_rows.append((mssv, ten_sv, diem))

        if not matched_any:
            messagebox.showwarning("Không tìm thấy",
                                   "❌ Không tìm thấy MSSV nào phù hợp trong ảnh bảng điểm.\nVui lòng kiểm tra lại ảnh.")
            return

        # === Tạo cửa sổ popup ===
        popup = ctk.CTkToplevel(self)
        popup.title(f"Chỉnh sửa điểm - {ten_cot_diem}")
        popup.geometry("600x400")
        popup.lift()
        popup.attributes("-topmost", True)

        frame = ctk.CTkFrame(popup)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        columns = ("MSSV", "Tên SV", "Điểm")
        tree = ttk.Treeview(frame, columns=columns, show="headings")
        tree.pack(fill="both", expand=True)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=180)

        for mssv, ten_sv, diem in sinh_vien_rows:
            tree.insert("", "end", values=(mssv, ten_sv, diem))

        # === Cho phép chỉnh sửa cột điểm ===
        def on_double_click(event):
            item_id = tree.identify_row(event.y)
            col = tree.identify_column(event.x)

            if col != "#3":
                return

            x, y, width, height = tree.bbox(item_id, column=col)
            value = tree.set(item_id, column=col)

            entry = ttk.Entry(tree)
            entry.place(x=x, y=y, width=width, height=height)
            entry.insert(0, value)
            entry.focus()

            def on_focus_out(event):
                new_value = entry.get()
                tree.set(item_id, column=col, value=new_value)
                entry.destroy()

            entry.bind("<FocusOut>", on_focus_out)
            entry.bind("<Return>", lambda e: on_focus_out(e))

        tree.bind("<Double-1>", on_double_click)

        # === Lưu vào CSDL ===
        def save_data():
            for item in tree.get_children():
                mssv, _, diem = tree.item(item, "values")
                if diem.strip() == "":
                    continue  # bỏ qua dòng chưa có điểm

                try:
                    diem_float = float(diem)
                    if diem_float < 0 or diem_float > 10:
                        raise ValueError
                except ValueError:
                    messagebox.showerror("Lỗi", f"Điểm không hợp lệ cho MSSV {mssv}. Hãy kiểm tra lại.")
                    return
                id_cot_diem = next(item['id'] for item in self.ds if item['ten_cot_diem'] == ten_cot_diem)
                self.controller.update(mssv, id_cot_diem, diem)

            popup.destroy()
            messagebox.showinfo("Thành công", "✅ Đã cập nhật điểm vào bảng điểm.")
            self.refresh_columns_and_data()

        save_btn = ctk.CTkButton(popup, text="Lưu dữ liệu vào bảng điểm", command=save_data)
        save_btn.pack(pady=10)

    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Lưu bảng điểm đơn giản"
        )
        if not file_path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Bảng Điểm"

        # Định nghĩa font Times New Roman, size 14
        default_font = Font(name="Times New Roman", size=14)
        bold_font = Font(name="Times New Roman", size=14, bold=True)

        # Căn giữa
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Định nghĩa đường viền
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Ghi tiêu đề cột
        headers = ["MSSV", "Tên Sinh Viên", "Điểm"]
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment

        # Ghi dữ liệu từng dòng
        for row_index, item in enumerate(self.tree.get_children(), start=2):
            values = self.tree.item(item, "values")
            try:
                mssv = int(values[0])  # ép kiểu để không bị cảnh báo số lưu dưới dạng text
            except:
                mssv = values[0]

            ho_ten = values[1]
            diem = ""  # để trống

            row_data = [mssv, ho_ten, diem]
            for col_index, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.font = default_font
                cell.alignment = center_alignment
                cell.border = thin_border

        # Auto fit độ rộng cột
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 4
            sheet.column_dimensions[column].width = adjusted_width

        try:
            workbook.save(file_path)
            messagebox.showinfo("Thành công", f"Đã xuất file Excel:\n{file_path}")
            os.startfile(file_path)  # Mở file Excel sau khi lưu
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu file Excel:\n{str(e)}")

    def edit_selected_row(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Chọn dòng", "Vui lòng chọn một sinh viên để sửa.")
            return

        values = self.tree.item(selected_item, "values")
        if not values:
            return

        # Tạo popup sửa điểm
        popup = ctk.CTkToplevel(self)
        popup.title("Sửa điểm sinh viên")
        popup.geometry("400x500")

        # Hiển thị MSSV (không cho sửa)
        ctk.CTkLabel(popup, text="MSSV").pack(anchor="w", padx=10, pady=2)
        ctk.CTkLabel(popup, text=values[0]).pack(anchor="w", padx=10, pady=2)

        # Hiển thị Họ tên (không cho sửa)
        ctk.CTkLabel(popup, text="Họ tên").pack(anchor="w", padx=10, pady=2)
        ctk.CTkLabel(popup, text=values[1]).pack(anchor="w", padx=10, pady=2)

        entry_vars = []
        for i, (ten_cot, _) in enumerate(self.cot_diem_list):
            ctk.CTkLabel(popup, text=ten_cot).pack(anchor="w", padx=10, pady=2)
            var = ctk.StringVar(value=values[i + 2])  # +2 vì MSSV và Họ tên đứng đầu
            entry = ctk.CTkEntry(popup, textvariable=var)
            entry.pack(fill="x", padx=10, pady=2)
            entry_vars.append(var)

        def save_edits():
            mssv = values[0]
            for i, var in enumerate(entry_vars):
                diem_str = var.get()
                try:
                    diem = float(diem_str)
                    if not (0 <= diem <= 10):
                        raise ValueError  # Kiểm tra khoảng điểm
                except ValueError:
                    messagebox.showerror("Lỗi", f"Điểm không hợp lệ cho {self.cot_diem_list[i][0]}. Vui lòng nhập số từ 0 đến 10.")
                    return  # Dừng lưu nếu có lỗi

                id_cot_diem = self.id_cot_diem[i][0]
                self.controller.update(mssv, id_cot_diem, diem)

            self.load_data()
            self.tinh_diem_kiem_tra()
            popup.destroy()
            messagebox.showinfo("Thành công", "Đã cập nhật điểm.")

        save_btn = ctk.CTkButton(popup, text="Lưu", command=save_edits)
        save_btn.pack(pady=10)

    def tinh_diem_kiem_tra(self):
        """Tính điểm kiểm tra (giả sử điểm thành phần luôn hợp lệ)."""
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            mssv = values[0]  # Lấy MSSV của sinh viên
            diem_thanh_phan = []
            try:
                for x in values[2:-1]:
                    if x != "":  # Kiểm tra ô điểm không rỗng
                        diem_thanh_phan.append(float(x))
                    else:
                        diem_thanh_phan.append(0.0)  # Xử lý khi ô điểm rỗng

                tong_trong_so = sum(ts for _, ts in self.cot_diem_list)
                if tong_trong_so == 0:
                    diem_kiem_tra = 0  # Xử lý khi tổng trọng số bằng 0
                else:
                    diem_kiem_tra = sum(
                        diem * trong_so for diem, (_, trong_so) in
                        zip(diem_thanh_phan, self.cot_diem_list)) / tong_trong_so

                self.tree.set(item, column="Điểm Kiểm Tra", value=round(diem_kiem_tra, 2))
                self.diemcontroller.insert_diem_kiem_tra(mssv, self.ma_lop,
                                                         diem_kiem_tra)  # Update điểm kiểm tra sau khi tính

            except ValueError:  # Xử lý lỗi khi điểm không phải là số
                messagebox.showerror("Lỗi", "Điểm thành phần không hợp lệ. Vui lòng nhập số.")



